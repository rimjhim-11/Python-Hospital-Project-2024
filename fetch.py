from openpyxl import load_workbook
from data import *
import tkinter

from tkinter import messagebox
from therapy import calculate, create_therapy_list

from updateScreen import *

patient_details = []
codes = []


patient_details = []

def get(code,  screen, option): #mobile,
    global patient_details
    if option == 6:
        data = code.get()
        code_temp_var = data[:4]
    else:
        code_temp_var = str(code)


    match_found = False
    patient_details = []
    wb = load_workbook('C:/Users/0034DK744/Documents/Development/Hospital Project/Hospital.xlsx')
    ws = wb["Master Sheet"]
    for row in range(ws.max_row+1):
        cell = str(ws.cell(row+1, 1).value)
        if cell in code_temp_var:
            match_found = True
            print('found a match for this code in the master sheet:' )
            #row_found = row+1
            for col in range(10):
                patient_details.append(ws.cell(row+1,col+1).value)
                #pass
            print(patient_details)
        if match_found:
            break

    print("Out of the for loop")
    print(f"Get Update Screen {patient_details}")
    get_update_screen(screen, patient_details)


def fetch_by_dist_widget(parent_screen):
    temp_screen = Toplevel(parent_screen)
    temp_screen.title("Codes and Mobile details")

    # specify size
    temp_screen.geometry("400x100")
    temp_screen.config(padx=20, pady=20, bg='lemon chiffon')
    temp_screen.grid_columnconfigure(2, weight=1)
    temp_screen.wm_transient(parent_screen)
    var_f_code = tkinter.StringVar()
    var_f_code.set('Code:Mobile ')
    codes_drop = OptionMenu(temp_screen, var_f_code , *codes)
    codes_drop.config(width=20, height= 1)
    codes_drop.grid(row=0, column=0)

    ok_button = Button(temp_screen, text='Ok',
                                   command=lambda: get(var_f_code, parent_screen, 6)) #,var_f_mob

    ok_button.config(width=20, height=1)
    ok_button.grid(row=2, column=3)

def fetch_from_sheet(search_key, option, parent_screen, local_data_list):
    global patient_details
    print("IN FETCH FROM SHEET FUNCTION FROM FETCH.py")
    match_found = False
    key = str(search_key.get())
    print(key)
    wb = load_workbook('C:/Users/0034DK744/Documents/Development/Hospital Project/Hospital.xlsx')
    ws = wb["Master Sheet"]
    if key == "":
        messagebox.showerror(title="ERROR!!", message="Enter a valid Code/Mobile/District")
    else:
        print(key)
        print("Entered option 1")
        for column in range (option-1,option):
            column += 1
            for row  in range (ws.max_row+ 1):
                row += 1
                cell = ws.cell(row, column)
                if cell.value == None:
                    cell.value = "none"
                if key.lower() in str(cell.value).lower():
                    match_found = True
                    if option in [1,8]:
                        code_found = ws.cell(row,1).value
                        messagebox.showinfo(title="Information", message="Found a match")
                        # match_found = True
                        break
                    else:
                        codes.append(f'{ws.cell(row, 1).value}:{ws.cell(row, 8).value}')
                # else:
                #     messagebox.showerror(title="CODE ERROR!!")

            if match_found: #and option in [1,6]:
                break
            else:
                messagebox.showerror(title="Error!!!", message="Code/Mobile/Dist not found")
        #else:

        print("Break")
        if option == 6 and match_found:
            fetch_by_dist_widget(parent_screen)
        elif match_found:
            get(code_found,parent_screen,option)


